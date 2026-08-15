# -*- coding: utf-8 -*-
"""
여비 지급명세서 엑셀 출력기 (openpyxl)
=====================================================================
HWP 양식 「여비 지급명세서」를 openpyxl 병합셀로 100% 재현.

양식 구조 (15열):
  1  출장자(소속/직급/성명)
  2  출장목적
  3  출장월일(시간)
  4  출발          ┐
  5  도착          │ 출장지(경로요금 포함)
  6  종별          │
  7  거리/요금     ┘
  8  식비
  9  숙박비
  10 일비
  11 현지 교통비
  12 기타
  13 계
  14 청구액및 수령액
  15 영수인(청구액외 포기함)

행 구조:
  · 헤더 2행 (병합)
  · 자차 운전자 = 4행 블록
      갈때행   : 출발지 / 최종도착 / 자가용 / [거리]Km  (유류비는 아래 겹행)
      올때행   : 최종도착 / 출발지 / 자가용 / [거리]Km
      통행료행 : 고속도로통행료 ×3 / 0Km / [통행료]
      합계행   : 공무용차량(미/사용) / 합계 / [총거리]Km / … / [운임+통행료]
    ※ 실제 양식은 갈때·올때·통행료의 '거리'와 '요금'이 한 칸 안에 위·아래로
       들어가 있어, 여기서는 거리행/요금행을 세로로 나누지 않고
       '거리/요금' 열(7열)에 "117Km" → 다음 겹 "18,200원" 식으로
       두 줄을 별도 행으로 배치한다. (양식 PrvText 순서와 동일)
  · 동승자 = 2행 블록 (운임·통행료 0, 경로 '-')
  · 맨 아래 총합계 1행

계산 규칙:
  · 유류비 = (편도거리 ÷ 표준연비) × 유가   ← 갈때·올때 각각
  · 일비 25,000×일수 (공무차 1/2), 식비 25,000×일수
  · 숙박비 상한×박수 (서울10만/광역시8만/그밖7만)
  · 십원 미만 버림, 공무차는 운임·통행료 0
  · days = 종료-시작+1, nights = 종료-시작
  · 경유지: 갈때 = 출발→경유지들→최종도착 (waypoints),
            올때 = 최종도착→출발 (직행)
"""

from __future__ import annotations
import io
import math
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================
# 상수 (기존 스트림릿 코드와 동일 — 재사용 목적으로 여기에도 정의)
# =============================================================
FUEL_STANDARDS = {
    "휘발유":            {"eff": 11.97, "unit": "km/L"},
    "경유":              {"eff": 12.52, "unit": "km/L"},
    "LPG":               {"eff": 8.83,  "unit": "km/L"},
    "하이브리드":         {"eff": 15.37, "unit": "km/L"},
    "플러그인하이브리드":  {"eff": 10.61, "unit": "km/L"},
    "전기":              {"eff": 5.22,  "unit": "km/kWh"},
}

ALLOWANCE = {
    "일비_1일": 25000,
    "식비_1일": 25000,
    "숙박비_상한_서울": 100000,
    "숙박비_상한_광역시": 80000,
    "숙박비_상한_그밖": 70000,
}

SUKBAK_KEY = {
    "서울": "숙박비_상한_서울",
    "광역시": "숙박비_상한_광역시",
    "그밖": "숙박비_상한_그밖",
}


def floor10(value) -> int:
    """십원 미만 버림."""
    try:
        return (int(value) // 10) * 10
    except (TypeError, ValueError):
        return 0


def calc_fuel_cost(distance_km: float, efficiency: float, unit_price: float) -> float:
    """유류비 = 거리 ÷ 연비 × 단가."""
    if efficiency <= 0:
        return 0.0
    return (distance_km / efficiency) * unit_price


# =============================================================
# 데이터 모델
# =============================================================
@dataclass
class Person:
    """명세서 한 사람(운전자 또는 동승자)."""
    소속: str = ""
    직급: str = ""
    성명: str = ""
    목적: str = ""
    시작일: Optional[date] = None
    종료일: Optional[date] = None
    출발지명: str = ""          # 표기용 (예: 경북도청)
    도착지명: str = ""          # 표기용 (예: 대구)
    경유지명: str = ""          # 표기용 (쉼표구분, 있으면 "대구, 경산")
    유종: str = "휘발유"
    is_driver: bool = True      # 운전자 여부
    use_official_car: bool = False  # 공무용 차량

    # 계산 입력값 (일괄처리에서 카카오/유가로 채움)
    갈때거리_km: float = 0.0    # 출발→(경유지)→최종도착
    올때거리_km: float = 0.0    # 최종도착→출발
    유가: float = 0.0           # 원/L 또는 원/kWh
    통행료_갈때: int = 0
    통행료_올때: int = 0
    숙박지역: str = "그밖"       # 서울/광역시/그밖

    # 산출 결과 (calc()가 채움)
    _r: dict = field(default_factory=dict)

    # ---- 파생값 ----
    @property
    def days(self) -> int:
        if not self.시작일 or not self.종료일:
            return 1
        return (self.종료일 - self.시작일).days + 1

    @property
    def nights(self) -> int:
        if not self.시작일 or not self.종료일:
            return 0
        return (self.종료일 - self.시작일).days

    @property
    def 소속직급성명(self) -> str:
        parts = [p for p in (self.소속, self.직급, self.성명) if p]
        return " ".join(parts)

    @property
    def 출장월일(self) -> str:
        if not self.시작일 or not self.종료일:
            return ""
        s, e = self.시작일, self.종료일
        return f"{s:%m-%d}~{e:%m-%d}"

    @property
    def 도착표기(self) -> str:
        """경유지 있으면 '최종도착(경유지 경유)'."""
        if self.경유지명:
            return f"{self.도착지명}({self.경유지명} 경유)"
        return self.도착지명

    def calc(self) -> dict:
        """여비 산출. 결과 dict를 self._r에 저장하고 반환."""
        eff = FUEL_STANDARDS.get(self.유종, FUEL_STANDARDS["휘발유"])["eff"]

        # --- 유류비 (자차·운전자만) ---
        if self.is_driver and not self.use_official_car:
            fuel_go = floor10(calc_fuel_cost(self.갈때거리_km, eff, self.유가))
            fuel_back = floor10(calc_fuel_cost(self.올때거리_km, eff, self.유가))
            toll_go = floor10(self.통행료_갈때)
            toll_back = floor10(self.통행료_올때)
        else:
            fuel_go = fuel_back = toll_go = toll_back = 0

        운임 = fuel_go + fuel_back
        통행료 = toll_go + toll_back

        # --- 일비 (공무차 1/2) ---
        ilbi = ALLOWANCE["일비_1일"] * self.days
        if self.use_official_car:
            ilbi = ilbi // 2
        ilbi = floor10(ilbi)

        # --- 식비 ---
        sikbi = floor10(ALLOWANCE["식비_1일"] * self.days)

        # --- 숙박비 ---
        cap = ALLOWANCE[SUKBAK_KEY.get(self.숙박지역, "숙박비_상한_그밖")]
        sukbak = floor10(cap * max(0, self.nights))

        현지교통 = 0
        기타 = 0

        계 = 운임 + 통행료 + sikbi + sukbak + ilbi + 현지교통 + 기타

        self._r = {
            "fuel_go": fuel_go, "fuel_back": fuel_back,
            "toll_go": toll_go, "toll_back": toll_back,
            "운임": 운임, "통행료": 통행료,
            "식비": sikbi, "숙박비": sukbak, "일비": ilbi,
            "현지교통": 현지교통, "기타": 기타, "계": 계,
            "총거리": round(self.갈때거리_km + self.올때거리_km, 0),
            "갈때거리": round(self.갈때거리_km, 0),
            "올때거리": round(self.올때거리_km, 0),
        }
        return self._r


@dataclass
class Group:
    """운전자 1명 + 동승자 N명 (같은 차)."""
    driver: Person
    passengers: list = field(default_factory=list)

    def all_people(self):
        return [self.driver] + list(self.passengers)


# =============================================================
# 엑셀 스타일
# =============================================================
_THIN = Side(style="thin", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
FONT = Font(name="맑은 고딕", size=9)
FONT_B = Font(name="맑은 고딕", size=9, bold=True)
FONT_TITLE = Font(name="맑은 고딕", size=16, bold=True)

NCOLS = 15


def _won(n) -> str:
    try:
        return f"{int(round(n)):,}원"
    except (TypeError, ValueError):
        return "0원"


def _km(n) -> str:
    try:
        return f"{int(round(n)):,}Km"
    except (TypeError, ValueError):
        return "0Km"


def _style_cell(ws, row, col, value, *, font=FONT, align=CENTER, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    c.border = BORDER
    if fill:
        c.fill = fill
    return c


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _apply_border_range(ws, r1, c1, r2, c2):
    """병합영역 전체 테두리 보정."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# =============================================================
# 명세서 작성
# =============================================================
def build_workbook(groups, *, dept_footer="경상북도 건설도시국 토지정보과",
                   signer="박00") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "여비지급명세서"

    # ---- 열 너비 ----
    widths = [16, 20, 14, 8, 8, 7, 9, 8, 8, 8, 8, 8, 10, 11, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ---- 제목 ----
    _merge(ws, row, 1, row, NCOLS)
    t = _style_cell(ws, row, 1, "여비 지급명세서", font=FONT_TITLE, align=CENTER)
    ws.row_dimensions[row].height = 30
    _apply_border_range(ws, row, 1, row, NCOLS)
    row += 1

    # ---- 헤더 2행 ----
    hdr_top = row
    hdr_bot = row + 1
    # 1열 출장자 (2행 병합 아님 — 아래에 소속/직급/성명·공무용차량 서브헤더)
    _merge(ws, hdr_top, 1, hdr_top, 1)
    _style_cell(ws, hdr_top, 1, "출장자", font=FONT_B, fill=HEADER_FILL)
    _style_cell(ws, hdr_bot, 1, "소속/직급/성명", font=FONT_B, fill=HEADER_FILL)

    # 2열 출장목적 (세로 2행 병합)
    _merge(ws, hdr_top, 2, hdr_bot, 2)
    _style_cell(ws, hdr_top, 2, "출장목적", font=FONT_B, fill=HEADER_FILL)

    # 3열 출장월일(시간)
    _merge(ws, hdr_top, 3, hdr_bot, 3)
    _style_cell(ws, hdr_top, 3, "출장월일\n(출장시간)", font=FONT_B, fill=HEADER_FILL)

    # 4~7열 출장지(경로요금 포함) — 상단 병합, 하단 출발/도착/종별/거리요금
    _merge(ws, hdr_top, 4, hdr_top, 7)
    _style_cell(ws, hdr_top, 4, "출장지(경로요금 포함)", font=FONT_B, fill=HEADER_FILL)
    for c, label in zip(range(4, 8), ["출발", "도착", "종별", "거리/요금"]):
        _style_cell(ws, hdr_bot, c, label, font=FONT_B, fill=HEADER_FILL)

    # 8~15열 금액항목 (세로 2행 병합)
    money_labels = ["식비", "숙박비", "일비", "현지 교통비", "기타", "계",
                    "청구액및\n수령액", "영수인\n(청구액외\n포기함)"]
    for c, label in zip(range(8, 16), money_labels):
        _merge(ws, hdr_top, c, hdr_bot, c)
        _style_cell(ws, hdr_top, c, label, font=FONT_B, fill=HEADER_FILL)

    _apply_border_range(ws, hdr_top, 1, hdr_bot, NCOLS)
    ws.row_dimensions[hdr_top].height = 20
    ws.row_dimensions[hdr_bot].height = 28
    row = hdr_bot + 1

    # ---- 총합계 누적용 ----
    tot = {"식비": 0, "숙박비": 0, "일비": 0, "현지교통": 0, "기타": 0,
           "계": 0, "운임": 0, "통행료": 0}

    # ---- 그룹별 블록 ----
    for g in groups:
        for p in g.all_people():
            p.calc()
            if p.is_driver:
                row = _write_driver_block(ws, row, p)
            else:
                row = _write_passenger_block(ws, row, p)
            r = p._r
            tot["식비"] += r["식비"]
            tot["숙박비"] += r["숙박비"]
            tot["일비"] += r["일비"]
            tot["현지교통"] += r["현지교통"]
            tot["기타"] += r["기타"]
            tot["계"] += r["계"]
            tot["운임"] += r["운임"]
            tot["통행료"] += r["통행료"]

    # ---- 총합계 행 ----
    row = _write_total_row(ws, row, tot, dept_footer, signer)

    return wb


def _write_driver_block(ws, row, p: Person) -> int:
    """자차 운전자 4행 블록. 반환: 다음 시작 행."""
    r = p._r
    r0 = row          # 갈때
    r1 = row + 1      # 올때
    r2 = row + 2      # 통행료
    r3 = row + 3      # 합계
    종별 = "공무용차량" if p.use_official_car else "자가용"

    # --- 1열 소속/직급/성명 (갈때~통행료 3행 병합) ---
    _merge(ws, r0, 1, r2, 1)
    _style_cell(ws, r0, 1, p.소속직급성명, align=LEFT)
    # 합계행 1열: 공무용차량(미사용)/(사용)
    _style_cell(ws, r3, 1, f"공무용차량({'사용' if p.use_official_car else '미사용'})")

    # --- 2열 출장목적 (갈때~통행료 3행 병합) ---
    _merge(ws, r0, 2, r2, 2)
    _style_cell(ws, r0, 2, p.목적, align=LEFT)
    _style_cell(ws, r3, 2, "", )  # 합계행 목적 빈칸

    # --- 3열 출장월일 (갈때~통행료 3행 병합) ---
    _merge(ws, r0, 3, r2, 3)
    dt = p.출장월일
    time_part = "(09:00 ~ 18:00)"
    _style_cell(ws, r0, 3, f"{dt}\n{time_part}" if dt else "")
    _style_cell(ws, r3, 3, "")  # 합계행

    # --- 4~7열 경로 ---
    # 갈때행: 출발지 / 최종도착 / 종별 / 거리
    _style_cell(ws, r0, 4, p.출발지명)
    _style_cell(ws, r0, 5, p.도착표기)
    _style_cell(ws, r0, 6, 종별)
    _style_cell(ws, r0, 7, f"{_km(r['갈때거리'])}\n{_won(r['fuel_go'])}", font=FONT)
    # 올때행: 최종도착 / 출발지 / 종별 / 거리
    _style_cell(ws, r1, 4, p.도착지명)
    _style_cell(ws, r1, 5, p.출발지명)
    _style_cell(ws, r1, 6, 종별)
    _style_cell(ws, r1, 7, f"{_km(r['올때거리'])}\n{_won(r['fuel_back'])}", font=FONT)
    # 통행료행: 고속도로통행료 ×3(출발·도착·종별) / 통행료
    _style_cell(ws, r2, 4, "고속도로통행료")
    _style_cell(ws, r2, 5, "고속도로통행료")
    _style_cell(ws, r2, 6, "고속도로통행료")
    _style_cell(ws, r2, 7, f"{_km(0)}\n{_won(r['통행료'])}", font=FONT)
    # 합계행: 합계 / 총거리 (4~6 병합 '합계', 7열 총거리)
    _merge(ws, r3, 4, r3, 6)
    _style_cell(ws, r3, 4, "합계", font=FONT_B)
    _style_cell(ws, r3, 7, _km(r["총거리"]), font=FONT_B)

    # --- 8~14열 금액 (갈때~통행료 3행 세로병합, 값은 갈때행에) ---
    money = [("식비", 8), ("숙박비", 9), ("일비", 10),
             ("현지교통", 11), ("기타", 12), ("계", 13)]
    for key, col in money:
        _merge(ws, r0, col, r2, col)
        _style_cell(ws, r0, col, _won(r[key]), align=RIGHT)
        _style_cell(ws, r3, col, "")  # 합계행 빈칸
    # 14열 청구액및수령액 = 계 (동일)
    _merge(ws, r0, 14, r2, 14)
    _style_cell(ws, r0, 14, _won(r["계"]), align=RIGHT)
    # 합계행 14열: 운임+통행료 합
    _style_cell(ws, r3, 14, _won(r["운임"] + r["통행료"]), align=RIGHT, font=FONT_B)

    # --- 15열 영수인 (갈때~합계 4행 병합) ---
    _merge(ws, r0, 15, r3, 15)
    _style_cell(ws, r0, 15, "")

    _apply_border_range(ws, r0, 1, r3, NCOLS)
    for rr in (r0, r1, r2, r3):
        ws.row_dimensions[rr].height = 22
    return r3 + 1


def _write_passenger_block(ws, row, p: Person) -> int:
    """동승자 2행 블록 (운임·통행료 0, 경로 '-'). 반환: 다음 시작 행."""
    r = p._r
    r0 = row      # 데이터
    r1 = row + 1  # 합계

    # 1열
    _style_cell(ws, r0, 1, p.소속직급성명, align=LEFT)
    _style_cell(ws, r1, 1, f"공무용차량({'사용' if p.use_official_car else '미사용'})")
    # 2열 목적
    _style_cell(ws, r0, 2, p.목적, align=LEFT)
    _style_cell(ws, r1, 2, "")
    # 3열 월일
    dt = p.출장월일
    _style_cell(ws, r0, 3, f"{dt}\n(09:00 ~ 18:00)" if dt else "")
    _style_cell(ws, r1, 3, "")
    # 4~7 경로: 출발 '-', 도착 표기, 종별 '-', 거리 0Km
    _style_cell(ws, r0, 4, "-")
    _style_cell(ws, r0, 5, p.도착표기)
    _style_cell(ws, r0, 6, "-")
    _style_cell(ws, r0, 7, _km(0))
    _merge(ws, r1, 4, r1, 6)
    _style_cell(ws, r1, 4, "합계", font=FONT_B)
    _style_cell(ws, r1, 7, _km(0), font=FONT_B)
    # 8~13 금액
    money = [("식비", 8), ("숙박비", 9), ("일비", 10),
             ("현지교통", 11), ("기타", 12), ("계", 13)]
    for key, col in money:
        _style_cell(ws, r0, col, _won(r[key]), align=RIGHT)
        _style_cell(ws, r1, col, "")
    # 14 청구액및수령액
    _style_cell(ws, r0, 14, _won(r["계"]), align=RIGHT)
    _style_cell(ws, r1, 14, _won(0), align=RIGHT, font=FONT_B)
    # 15 영수인
    _merge(ws, r0, 15, r1, 15)
    _style_cell(ws, r0, 15, "")

    _apply_border_range(ws, r0, 1, r1, NCOLS)
    ws.row_dimensions[r0].height = 22
    ws.row_dimensions[r1].height = 22
    return r1 + 1


def _write_total_row(ws, row, tot, dept_footer, signer) -> int:
    """맨 아래 총합계 행."""
    r = row
    # 원본 양식: '합계'(1~7열 병합) | 운임+통행료 | 식비 | 숙박 | 일비 | 현지교통 | 기타 | 계 | 청구액
    _merge(ws, r, 1, r, 7)
    _style_cell(ws, r, 1, "합            계", font=FONT_B)
    운임통행 = tot["운임"] + tot["통행료"]
    _style_cell(ws, r, 8, _won(운임통행), align=RIGHT, font=FONT_B)      # 운임+통행료 합
    _style_cell(ws, r, 9, _won(tot["식비"]), align=RIGHT, font=FONT_B)
    _style_cell(ws, r, 10, _won(tot["숙박비"]), align=RIGHT, font=FONT_B)
    _style_cell(ws, r, 11, _won(tot["일비"]), align=RIGHT, font=FONT_B)
    _style_cell(ws, r, 12, _won(tot["현지교통"]), align=RIGHT, font=FONT_B)
    _style_cell(ws, r, 13, _won(tot["기타"]), align=RIGHT, font=FONT_B)
    _style_cell(ws, r, 14, _won(tot["계"]), align=RIGHT, font=FONT_B)
    _style_cell(ws, r, 15, _won(tot["계"]), align=RIGHT, font=FONT_B)
    _apply_border_range(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 22
    r += 1

    # 하단 서명/부서 행 (양식 마지막 줄)
    stamp = f"{signer} {datetime.now():%Y년 %m월 %d일 %H시 %M분 %S초}"
    _merge(ws, r, 1, r, 9)
    _style_cell(ws, r, 1, dept_footer, align=CENTER, font=FONT_B)
    _merge(ws, r, 10, r, 11)
    _style_cell(ws, r, 10, "1/1", align=CENTER)
    _merge(ws, r, 12, r, 15)
    _style_cell(ws, r, 12, stamp, align=CENTER)
    _apply_border_range(ws, r, 1, r, NCOLS)
    ws.row_dimensions[r].height = 22
    return r + 1


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================
# 카카오 경유지 지원 (기존 get_road_distance 확장판)
# =============================================================
def get_road_distance_waypoints(origin_xy, dest_xy, kakao_key, waypoints_xy=None):
    """
    좌표 → 도로 주행거리(m), 통행료(원). 경유지(waypoints) 지원.
    waypoints_xy: [(x,y), ...] 또는 None.
    카카오 mobility directions는 waypoints 파라미터 지원.
    """
    import requests
    KAKAO_DIRECTIONS_URL = "https://apis-navi.kakaomobility.com/v1/directions"
    if not kakao_key:
        return {"ok": False, "reason": "카카오 키 없음"}
    if not origin_xy or not dest_xy:
        return {"ok": False, "reason": "좌표 없음"}
    headers = {"Authorization": f"KakaoAK {kakao_key}"}
    params = {
        "origin": f"{origin_xy[0]},{origin_xy[1]}",
        "destination": f"{dest_xy[0]},{dest_xy[1]}",
        "priority": "RECOMMEND",
        "car_fuel": "GASOLINE",
        "car_hipass": "false",
    }
    if waypoints_xy:
        params["waypoints"] = "|".join(f"{x},{y}" for x, y in waypoints_xy)
    try:
        r = requests.get(KAKAO_DIRECTIONS_URL, headers=headers, params=params, timeout=15)
        if r.status_code in (401, 403):
            return {"ok": False, "reason": "길찾기 권한 미승인"}
        r.raise_for_status()
        summary = r.json()["routes"][0]["summary"]
        return {
            "ok": True,
            "distance_m": summary["distance"],
            "duration_s": summary["duration"],
            "toll": summary.get("fare", {}).get("toll", 0),
        }
    except Exception as e:
        return {"ok": False, "reason": f"길찾기 오류: {e}"}
